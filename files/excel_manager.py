#!/usr/bin/python 
# -*- coding: utf-8 -*-
# @Author: 丁珂
# @Time: 2019/7/31 15:04
import datetime
from flask import flash
# sys.path.append()
import xlrd
import os
import re
from openpyxl import load_workbook

import my_path
from my_log import debug_logger, static_logger


def remove_white_space(data, flag=0):
    '''
    取掉字符串中的空格
    :return:
    '''
    # 表示传入的是号码
    new_data = []
    if flag:
        # 进行合法性检查
        for index, number in enumerate(data):
            number = str(number).strip()
            number = number.replace(' ', '')
            # 检查号码中数字长度是否为11位 且等于总长度11(说明没有其他字符)
            # print(1, number, type(number))
            try:
                if number is not None and len(number) == len(re.search('\d+', number).group()) == 11:
                    new_data.append(number)
                else:
                    msg = 'Excel中第' + str(index + 2) + '行' + '电话号码有误' + '号码为:' + number
                    # print(2,number+str(1))
                    static_logger.error(msg)
                    flash(msg, 'EXCEL_ERROR')
            except:
                msg = 'Excel中第' + str(index + 2) + '行' + '电话号码有误' + '号码为:' + number
                # print(2,number+str(1))
                static_logger.error(msg)
                flash(msg, 'EXCEL_ERROR')
        return new_data
    # 其余字段不进行合法性检查
    else:
        for i in data:
            i = str(i)
            new_data.append(i.replace(' ', ''))
        return new_data


def get_phone_numbers(sheet):
    '''
    遍历电话号码 判断如果是小数就转换为整数
    :param sheet:
    :return: 整数类型的电话号码list
    :type: list
    '''
    phone_numbers = []
    # 遍历电话号码
    nrows = sheet.nrows
    # 电话号码在表格中的列的位置
    col = 3
    for i in range(1, nrows):
        # 数据类型
        ctype = sheet.cell(i, col).ctype
        # 值
        value = sheet.cell_value(i, col)
        # 如果是小数
        if ctype == 2 and value % 1 == 0.0:
            # 转换成整数
            value = int(value)
        phone_numbers.append(value)
    return phone_numbers


def change_to_dict(file_path):
    '''
    将给定的Excel表格中需要的数据提取出来转换成dict返回
    :param file_path:
    :return: dict格式的数据
    '''
    # 读取文件
    try:
        workbook = xlrd.open_workbook(filename=file_path)
    except FileNotFoundError:
        static_logger.error('上传的Excel文件未找到')
        debug_logger.debug('上传的Excel文件未找到')

    # 选择第一个sheet
    sheet = workbook.sheet_by_index(0)
    # 构建dict
    D = dict.fromkeys(['name', 'express_number', 'phone_number'])
    D['name'] = remove_white_space(sheet.col_values(1)[1:])
    # 快递号码
    D['express_number'] = remove_white_space(sheet.col_values(2)[1:])
    # 电话
    D['phone_number'] = remove_white_space(get_phone_numbers(sheet), 1)
    # 看数据有无异常
    if len(D['name']) == len(D['express_number']) == len(D['phone_number']):
        return D
    else:
        static_logger.error('导入Excel数据异常')
        debug_logger.debug('导入Excel数据异常')
        return 1



def write(data, username):
    # 检测用户对应的文件名是否存在
    file_name = str(datetime.date.today()) + username + '.xlsx'
    file_path = os.path.join(my_path.EXCELS_PATH, 'download', file_name)
    # 存在对应文件则载入
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        # 不存在从模板获取workbook
        workbook = select_template_as_file()
    #
    if workbook and workbook != 1:
        # grab the active worksheet
        sheet = workbook.active
        sheet.append(data)
        # next_row = sheet.max_row + 1
        # for i in range(1,len(data)+1):
        #     sheet.cell(next_row, i).value = data[i-1]
        # Save the file
        try:
            workbook.save(os.path.join(my_path.EXCELS_PATH, 'download', file_name))
        except:
            flash('Excel表格被其他程序占用,请先关闭excel表格', 'EXCEL_ERROR')
            return 1
        # row = [u'收件人姓名', u'手机/电话', u'省', u'市', u'区', u'地址', u'物品名称', u'备注', u'数量', u'销售金额']
        return 0
    return 1


def select_file_as_template(file_path):
    '''
    载入Excel表格作为使用的模板
    :param file_path: 指定的excel表格的路径
    :return:
    '''
    # 加载源文件

    workbook = load_workbook(file_path)
    # 提取没有后缀名的文件路径
    file_path_without_ext = file_path.rsplit('.', 1)[0]
    debug_logger.debug(file_path_without_ext)
    # 保存为xltx格式的模板文件
    workbook.save(file_path_without_ext + '.xltx')
    static_logger.error('读取' + file_path + '失败')

    return 0


def select_template_as_file():
    '''
    返回从模板新创建的workbook对像
    :return: workbook
    :type: openpyxl.worksheet成功; 1失败
    '''
    workbook = 1
    try:
        workbook = load_workbook(os.path.join(my_path.ProjectPath, 'templates', 'template.xltx'))
        sheet = workbook.active
        workbook.template = False
    except:
        debug_logger.debug('使用Excel模板文件失败')
        flash('使用Excel模板文件失败', 'EXCEL_ERROR')
    return workbook


def test(x, y, *data):
    print(data)
    print(x, y)


if __name__ == '__main__':
    # write(os.path.join('..\Excels\download', str(datetime.date.today()) + 'username' + '.xlsx'))

    # use_template(os.path.join('..\Excels\download', str(datetime.date.today()) + 'username' + '.xltx'))

    # select_file_as_template(r'C:\Users\Administrator\PycharmProjects\test1\dadago\web_by_flask\templates\template.xlsx')
    # select_template_as_file().save('t.xlsx')
    write(['1', '2', '3'], 'test')
